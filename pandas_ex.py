import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime

pd.options.display.max_columns = None

NEW_FILE_PATH = "fosslight_report_src_241129_1739.xlsx"
OLD_FILE_PATH = "fosslight_report_src_241129_1144.xlsx"

FILE_PATH = NEW_FILE_PATH

def save_prettier(data_fr, sheet_title):
    now = datetime.now()
    file_name = "results_" + now.strftime("%Y%m%d_%H%M") + ".xlsx"

    # 열 너비 자동 조정
    data_fr.to_excel(file_name)
    wb = load_workbook(file_name)
    ws = wb.active
    ws.title = sheet_title

# 1. 컬럼 크기를 내용의 최대치를 기준으로 하기.
    # for col in ws.columns:
    #     max_length = 0
    #     col_letter = get_column_letter(col[0].column)
    #     for cell in col:
    #         try:
    #             max_length = max(max_length, len(str(cell.value)))
    #         except:
    #             pass
    #     adjusted_width = max_length + 2  # 약간 여유 추가
    #     ws.column_dimensions[col_letter].width = adjusted_width
# 2. 컬럼의 크기를 딕셔너리로 설정
    column_widths = {
        'A': 5,  # A열의 너비를 20으로 설정
        'B': 5,  # B열의 너비를 30으로 설정
        'C': 100,  # C열의 너비를 15으로 설정
        'D': 20,
        'E': 20,
        'F': 20,
        'H': 20,
        'I': 20,
        'J': 20,
        'K': 50,
        'L': 50,
    }
    # 각 열의 너비 설정
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(file_name)


def check_diff(new_file, old_file):
    df_old = pd.read_excel(OLD_FILE_PATH, sheet_name="SRC_FL_Source", index_col=None)
    df_new = pd.read_excel(NEW_FILE_PATH, sheet_name="SRC_FL_Source", index_col=None)
    cond1 = df_new['Source Path'].isin(df_old['Source Path'])
    df_new_only = df_new[~cond1]
    df_new_only.to_excel("df_new_only.xlsx")
    print(df_old.head())


def make_df_from_excel(file_path, sheet):
    df = pd.read_excel(file_path, sheet_name=sheet)

    return df


def exe_rule(df):
    # Rule1 적용 -
    # 정규표현식을 사용한 검색
    # pattern = r'^b.*@example\.com'  # 'b'로 시작하고 '@example.com'으로 끝나는 이메일
    # filtered_df = df[df['email'].str.contains(pattern, regex=True)]

    # r'\.(D|dep)$' - .D나 .dep로 끝나는 문자열
    pattern1 = r'\.(D|dep|dependencies|txt)$'

    cond1 = df['Source Path'].str.contains('L3UESIM')
    cond2 = df['Source Path'].str.contains('third_party')
    cond3 = df['Source Path'].str.contains(pattern1, regex=True)

    # print(df.loc[cond1])
    df.loc[(cond1 | cond2 | cond3), 'Exclude'] = "Exclude"

    return df


def main(file_path):
    df = make_df_from_excel(FILE_PATH, "SRC_FL_Source")
    return df


if __name__ == "__main__":
    # check_diff(NEW_FILE_PATH, OLD_FILE_PATH)
    df = main(FILE_PATH)
    df = exe_rule(df)
    save_prettier(df, "SRC_FL_Source")

    filter1 = (df["Exclude"] != "Exclude")
    print(len(df.loc[filter1, "OSS Name"].value_counts()))
